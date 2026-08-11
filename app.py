import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime, timedelta
import math
import gc
import os
import sys

# --- MEMORY OPTIMIZATION ---
pd.options.mode.chained_assignment = None
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', 100)

# Force garbage collection at startup
gc.collect()

# --- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="Automated Multi-Site RAAWA Generator", 
    page_icon="📄", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- CUSTOM CSS FOR PROFESSIONAL DESIGN ---
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 2rem;
    }
    .custom-card {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .feature-box {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        margin: 0.5rem 0;
        border-left: 4px solid #667eea;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    .advantage-box {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        margin: 0.5rem 0;
        color: white;
    }
    .sidebar-content {
        padding: 1rem;
    }
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 0.5rem 2rem;
        border-radius: 5px;
        transition: all 0.3s ease;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }
    .stSuccess {
        background: linear-gradient(135deg, #84fab0 0%, #8fd3f4 100%);
        padding: 1rem;
        border-radius: 10px;
    }
    .info-box {
        background: #e3f2fd;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #2196f3;
    }
    .stDateInput {
        width: 100%;
    }
</style>
""", unsafe_allow_html=True)

# --- UTILITY FUNCTIONS ---
def format_contact_number(contact):
    """Format contact number to start with 0 and have 11 digits"""
    if not contact or pd.isna(contact) or contact == 'N/A':
        return "N/A"
    
    contact_str = str(contact).strip()
    contact_str = ''.join(filter(str.isdigit, contact_str))
    
    if len(contact_str) == 10:
        if contact_str.startswith('9'):
            return '0' + contact_str
        else:
            return contact_str
    elif len(contact_str) == 11:
        if contact_str.startswith('0'):
            return contact_str
        elif contact_str.startswith('63'):
            return '0' + contact_str[2:]
        else:
            return contact_str
    elif len(contact_str) == 12 and contact_str.startswith('639'):
        return '0' + contact_str[2:]
    else:
        return contact_str

def format_id_number(id_num):
    """Format ID number to remove decimal places"""
    if not id_num or pd.isna(id_num) or id_num == 'N/A':
        return "N/A"
    
    id_str = str(id_num).strip()
    
    try:
        if '.' in id_str:
            float_val = float(id_str)
            if float_val.is_integer():
                return str(int(float_val))
            else:
                return id_str
        else:
            return id_str
    except:
        return id_str

# --- FILE CHECK ---
def check_required_files():
    """Check if all required Excel files exist"""
    required_files = [
        "Globe FO Engr Conatct_Vendor.xlsx",
        "Requisitioner.xlsx",
        "EngrTech.xlsx",
        "MIN591__MANUAL RAAWA_APPLICATION_June8,2026.xlsx",
        "VIS SKSK Database-03.12.2026.xlsx",
        "ROGMA Unified SKSK DB for Sharing (1).xlsx"
    ]
    
    missing_files = []
    for file in required_files:
        if not os.path.exists(file):
            missing_files.append(file)
    
    if missing_files:
        st.error(f"❌ Missing required files: {', '.join(missing_files)}")
        st.info("Please upload all required Excel files to the app directory.")
        return False
    return True

# --- DATABASE LOADING ---
@st.cache_data
def load_databases():
    try:
        # --- LOAD SITE DATABASES FOR ALL REGIONS ---
        
        # 1. Load Mindanao (MIN) Database
        try:
            df_min = pd.read_excel(
                "Globe FO Engr Conatct_Vendor.xlsx", 
                sheet_name="MIN",
                dtype=str
            )
            df_min['REGION'] = 'MIN'
            st.info(f"✅ Loaded MIN database: {len(df_min)} sites")
        except Exception as e:
            st.warning(f"MIN database not found: {e}")
            df_min = pd.DataFrame()
        
        # 2. Load Visayas (VIS) Database with specific column mapping
        try:
            df_vis = pd.read_excel(
                "VIS SKSK Database-03.12.2026.xlsx",
                dtype=str
            )
            
            # Map VIS columns to standard format
            if 'PLAID' in df_vis.columns:
                df_vis['PLAID'] = df_vis['PLAID']
            else:
                df_vis['PLAID'] = df_vis.index.astype(str)
            
            if 'SITENAME' in df_vis.columns:
                df_vis['SITE'] = df_vis['SITENAME']
            else:
                df_vis['SITE'] = 'Unknown Site'
            
            if 'Site Address' in df_vis.columns:
                df_vis['SITE_ADD'] = df_vis['Site Address']
            elif 'SITE ADDRESS' in df_vis.columns:
                df_vis['SITE_ADD'] = df_vis['SITE ADDRESS']
            else:
                df_vis['SITE_ADD'] = 'N/A'
            
            if 'Super FO' in df_vis.columns:
                df_vis['NEW ENGINEER_AH'] = df_vis['Super FO']
            else:
                df_vis['NEW ENGINEER_AH'] = 'Unassigned'
            
            if 'TERRITORY' not in df_vis.columns:
                df_vis['TERRITORY'] = '0'
            
            if 'CONTACT NUMBER' not in df_vis.columns:
                df_vis['CONTACT NUMBER'] = 'N/A'
            
            df_vis['REGION'] = 'VIS'
            st.info(f"✅ Loaded VIS database: {len(df_vis)} sites")
        except Exception as e:
            st.warning(f"VIS database not found or error loading: {e}")
            df_vis = pd.DataFrame()
        
        # 3. Load Luzon (LUZ) Database
        try:
            df_luz = pd.read_excel(
                "ROGMA Unified SKSK DB for Sharing (1).xlsx",
                dtype=str
            )
            
            # Map LUZ columns to standard format
            if 'PLAID' in df_luz.columns:
                df_luz['PLAID'] = df_luz['PLAID']
            else:
                df_luz['PLAID'] = df_luz.index.astype(str)
            
            if 'SITENAME' in df_luz.columns:
                df_luz['SITE'] = df_luz['SITENAME']
            else:
                df_luz['SITE'] = 'Unknown Site'
            
            if 'SITE_ADD' in df_luz.columns:
                df_luz['SITE_ADD'] = df_luz['SITE_ADD']
            else:
                df_luz['SITE_ADD'] = 'N/A'
            
            if 'ENGINEER_AH' in df_luz.columns:
                df_luz['NEW ENGINEER_AH'] = df_luz['ENGINEER_AH']
            else:
                df_luz['NEW ENGINEER_AH'] = 'Unassigned'
            
            if 'TERRITORY' in df_luz.columns:
                df_luz['TERRITORY'] = df_luz['TERRITORY']
            else:
                df_luz['TERRITORY'] = '0'
            
            if 'CONTACT NUMBER' not in df_luz.columns:
                df_luz['CONTACT NUMBER'] = 'N/A'
            
            df_luz['REGION'] = 'LUZ'
            st.info(f"✅ Loaded LUZ database: {len(df_luz)} sites")
        except Exception as e:
            st.warning(f"LUZ database not found or error loading: {e}")
            df_luz = pd.DataFrame()
        
        # Combine all site databases
        df_sites = pd.concat([df_min, df_vis, df_luz], ignore_index=True)
        
        # If no sites loaded, create empty dataframe with required columns
        if df_sites.empty:
            st.error("No site databases loaded! Please check your files.")
            df_sites = pd.DataFrame(columns=['PLAID', 'SITE', 'REGION', 'TERRITORY', 'NEW ENGINEER_AH', 'CONTACT NUMBER', 'SITE_ADD'])
            return df_sites, None, None
        
        # Ensure all required columns exist
        required_cols = ['PLAID', 'SITE', 'REGION', 'TERRITORY', 'NEW ENGINEER_AH', 'CONTACT NUMBER', 'SITE_ADD']
        for col in required_cols:
            if col not in df_sites.columns:
                df_sites[col] = 'N/A'
        
        # Clean up data - only if columns exist
        if 'PLAID' in df_sites.columns:
            df_sites['PLAID'] = df_sites['PLAID'].astype(str).str.strip()
        if 'SITE' in df_sites.columns:
            df_sites['SITE'] = df_sites['SITE'].astype(str).str.strip()
        if 'TERRITORY' in df_sites.columns:
            df_sites['TERRITORY'] = df_sites['TERRITORY'].astype(str).str.replace('Territory', '', case=False).str.strip()
        if 'REGION' in df_sites.columns:
            df_sites['REGION'] = df_sites['REGION'].astype(str).str.upper().str.strip()
        if 'NEW ENGINEER_AH' in df_sites.columns:
            df_sites['NEW ENGINEER_AH'] = df_sites['NEW ENGINEER_AH'].astype(str).str.strip()
        if 'CONTACT NUMBER' in df_sites.columns:
            df_sites['CONTACT NUMBER'] = df_sites['CONTACT NUMBER'].astype(str).str.strip()
        if 'SITE_ADD' in df_sites.columns:
            df_sites['SITE_ADD'] = df_sites['SITE_ADD'].astype(str).str.strip()
        
        # Load Requisitioner Database with memory optimization
        try:
            df_req = pd.read_excel(
                "Requisitioner.xlsx", 
                header=1, 
                dtype=str
            )
            df_req.columns = df_req.columns.str.strip()
            
            # Show columns for debug
            st.info(f"Requisitioner columns: {', '.join(df_req.columns.tolist())}")
            
            if 'Territory no.' in df_req.columns:
                df_req['Territory no.'] = df_req['Territory no.'].astype(str).str.replace('Territory', '', case=False).str.strip()
            
            if 'Region' in df_req.columns:
                df_req['Region'] = df_req['Region'].astype(str).str.upper().str.strip()
                df_req['Region'] = df_req['Region'].apply(lambda x: x if x in ['LUZ', 'VIS', 'MIN'] else 'MIN')
            
            if 'Project' in df_req.columns:
                df_req['Project'] = df_req['Project'].astype(str).str.strip()
            
            # Batch process formatting for speed
            if 'ID #' in df_req.columns:
                df_req['ID #'] = df_req['ID #'].apply(lambda x: format_id_number(x) if pd.notna(x) else 'N/A')
            
            if 'Contact No.' in df_req.columns:
                df_req['Contact No.'] = df_req['Contact No.'].apply(lambda x: format_contact_number(x) if pd.notna(x) else 'N/A')
        except Exception as e:
            st.warning(f"Requisitioner.xlsx error: {e}")
            df_req = pd.DataFrame()
        
        # --- FIXED: Load Engineer/Technician Database ---
        try:
            df_engr_tech = pd.read_excel(
                "EngrTech.xlsx", 
                header=1,  # Row 2 contains column headers
                dtype=str
            )
            df_engr_tech.columns = df_engr_tech.columns.str.strip()
            
            # Debug: Show columns found
            st.info(f"EngrTech columns: {', '.join(df_engr_tech.columns.tolist())}")
            
            # Ensure we have the required columns
            # If the columns are not named correctly, map them
            if 'Name' not in df_engr_tech.columns:
                # Try to find the Name column
                for col in df_engr_tech.columns:
                    if col.lower() in ['name', 'full name', 'engineer']:
                        df_engr_tech.rename(columns={col: 'Name'}, inplace=True)
                        break
            
            if 'Company' not in df_engr_tech.columns:
                for col in df_engr_tech.columns:
                    if col.lower() in ['company', 'vendor', 'firm']:
                        df_engr_tech.rename(columns={col: 'Company'}, inplace=True)
                        break
            
            if 'ID No' not in df_engr_tech.columns:
                for col in df_engr_tech.columns:
                    if col.lower() in ['id no', 'id', 'sec id', 'employee id']:
                        df_engr_tech.rename(columns={col: 'ID No'}, inplace=True)
                        break
            
            if 'Region' not in df_engr_tech.columns:
                for col in df_engr_tech.columns:
                    if col.lower() in ['region', 'area']:
                        df_engr_tech.rename(columns={col: 'Region'}, inplace=True)
                        break
            
            # Format IDs
            if 'ID No' in df_engr_tech.columns:
                df_engr_tech['ID No'] = df_engr_tech['ID No'].astype(str).apply(format_id_number)
            
            # Clean up
            if 'Name' in df_engr_tech.columns:
                df_engr_tech = df_engr_tech.dropna(subset=['Name'], how='all')
                df_engr_tech = df_engr_tech[df_engr_tech['Name'].notna()]
                df_engr_tech = df_engr_tech[df_engr_tech['Name'].astype(str).str.strip() != '']
            
            # Fill NaN values
            for col in ['Company', 'ID No', 'Region']:
                if col in df_engr_tech.columns:
                    df_engr_tech[col] = df_engr_tech[col].fillna('').astype(str)
                else:
                    df_engr_tech[col] = ''
            
            # Clean Region - convert to uppercase and ensure valid values
            if 'Region' in df_engr_tech.columns:
                df_engr_tech['Region'] = df_engr_tech['Region'].str.upper().str.strip()
                # If Region is empty or not in LUZ/VIS/MIN, set to 'N/A'
                df_engr_tech['Region'] = df_engr_tech['Region'].apply(
                    lambda x: x if x in ['LUZ', 'VIS', 'MIN'] else 'N/A'
                )
            
            # Debug: Show data info
            st.info(f"EngrTech data loaded: {len(df_engr_tech)} rows")
            st.info(f"Regions found: {df_engr_tech['Region'].unique().tolist()}")
            st.info(f"Companies found: {df_engr_tech['Company'].unique().tolist()}")
                
        except Exception as e:
            st.warning(f"EngrTech.xlsx error: {e}")
            import traceback
            st.error(traceback.format_exc())
            df_engr_tech = pd.DataFrame(columns=['Name', 'Company', 'ID No', 'Region'])
        
        # Force garbage collection
        gc.collect()
        
        return df_sites, df_req, df_engr_tech
        
    except Exception as e:
        st.error(f"Error loading database files: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None, None, None

# --- RAAWA FILE CREATION ---
def create_raawa_file(matching_sites, personnel_list, scope_of_work, start_date, end_date, req_profile, facility_manager, batch_num=1, total_batches=1):
    """Helper function to create a single RAAWA file with dynamic font sizing"""
    try:
        template_file = "MIN591__MANUAL RAAWA_APPLICATION_June8,2026.xlsx"
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active 
        
        # Helper function to safely set a cell value - handles merged cells
        def safe_set_cell(row, col, value):
            """Safely set a cell value, handling merged cells"""
            cell = ws.cell(row=row, column=col)
            # Check if this cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Only write if this is the top-left cell of the merged range
                    if cell.coordinate == merged_range.start_cell.coordinate:
                        cell.value = value
                        return cell
                    else:
                        # Skip - this is a merged cell that's not the top-left
                        return None
            # Not merged, safe to write
            cell.value = value
            return cell
        
        # Helper function to safely set font on a cell
        def safe_set_font(row, col, font):
            """Safely set font on a cell, handling merged cells"""
            cell = ws.cell(row=row, column=col)
            # Check if this cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Only set font if this is the top-left cell of the merged range
                    if cell.coordinate == merged_range.start_cell.coordinate:
                        cell.font = font
                        return True
                    else:
                        return False
            # Not merged, safe to set font
            cell.font = font
            return True
        
        # --- SET REQUISITIONER DETAILS ---
        safe_set_cell(3, 4, req_profile["name"])   # D3
        safe_set_cell(4, 4, req_profile["dept"])   # D4
        
        id_value = format_id_number(req_profile["id"])
        safe_set_cell(4, 7, id_value)              # G4
        
        contact_value = format_contact_number(req_profile["contact"])
        safe_set_cell(4, 10, contact_value)        # J4
        
        # --- SET SITES ---
        base_site_row = 6
        num_sites = len(matching_sites)
        for idx, (_, row) in enumerate(matching_sites.iterrows()):
            curr_row = base_site_row + idx
            safe_set_cell(curr_row, 1, f"{row.get('PLAID', '')} - {row.get('SITE', '')}")
            safe_set_cell(curr_row, 4, row.get("SITE_ADD", "N/A"))
        
        # Hide unused site rows
        for r in range(base_site_row + num_sites, 16):
            ws.row_dimensions[r].hidden = True
        
        # --- SET DATES ---
        safe_set_cell(17, 4, start_date.strftime("%Y-%m-%d"))  # D17
        safe_set_cell(17, 5, end_date.strftime("%Y-%m-%d"))    # E17
        
        start_personnel_row = 19
        
        # --- SET PERSONNEL WITH MERGED CELL HANDLING ---
        for idx, person in enumerate(personnel_list):
            row_index = start_personnel_row + (idx // 2)
            col_offset = 0 if idx % 2 == 0 else 5
            
            formatted_id = format_id_number(person["id_no"])
            
            # Try to set each cell with merged cell handling
            # Name - Column 1 or 6
            name_cell = safe_set_cell(row_index, 1 + col_offset, person["name"])
            if name_cell:
                name_cell.font = Font(name="Calibri", size=6)
            
            # Company - Column 4 or 9
            company_cell = safe_set_cell(row_index, 4 + col_offset, person["company"])
            if company_cell:
                company_cell.font = Font(name="Calibri", size=6)
            
            # ID - Column 5 or 10
            id_cell = safe_set_cell(row_index, 5 + col_offset, formatted_id)
            if id_cell:
                id_cell.font = Font(name="Calibri", size=6)
        
        # Hide unused personnel rows
        for r in range(start_personnel_row + (len(personnel_list)//2 + 1), 39):
            ws.row_dimensions[r].hidden = True
        
        # --- SET SCOPE OF WORK ---
        if total_batches > 1:
            safe_set_cell(41, 1, f"{scope_of_work}\n\n(Page {batch_num} of {total_batches} for this location group)")
        else:
            safe_set_cell(41, 1, scope_of_work)
        
        # --- SET FACILITY MANAGER ---
        current_a48 = ws["A48"].value
        if current_a48:
            if "\n" in str(current_a48):
                parts = str(current_a48).split("\n", 1)
                if len(parts) == 2:
                    signature_part = parts[1]
                    safe_set_cell(48, 1, f"{facility_manager}\n{signature_part}")
                else:
                    safe_set_cell(48, 1, f"{facility_manager}\nSignature Over Printed Name / Date")
            else:
                safe_set_cell(48, 1, f"{facility_manager}\nSignature Over Printed Name / Date")
        else:
            safe_set_cell(48, 1, f"{facility_manager}\nSignature Over Printed Name / Date")
        
        # --- SET SECURITY APPROVER ---
        region = ''
        if not matching_sites.empty:
            first_site = matching_sites.iloc[0]
            region = str(first_site.get('REGION', '')).upper().strip()
        
        current_a50 = ws["A50"].value
        
        if region == 'VIS':
            if current_a50 and "\n" in str(current_a50):
                parts = str(current_a50).split("\n", 1)
                if len(parts) == 2:
                    signature_part = parts[1]
                    safe_set_cell(50, 1, f"JOJO A. VIRAY\n{signature_part}")
                else:
                    safe_set_cell(50, 1, f"JOJO A. VIRAY\nSignature Over Printed Name / Date")
            else:
                safe_set_cell(50, 1, f"JOJO A. VIRAY\nSignature Over Printed Name / Date")
            st.info(f"🔒 VIS Region detected - Security Approver set to: JOJO A. VIRAY")
        elif region == 'LUZ':
            if current_a50 and "\n" in str(current_a50):
                parts = str(current_a50).split("\n", 1)
                if len(parts) == 2:
                    signature_part = parts[1]
                    safe_set_cell(50, 1, f"TBD - Security Approver\n{signature_part}")
                else:
                    safe_set_cell(50, 1, f"TBD - Security Approver\nSignature Over Printed Name / Date")
            else:
                safe_set_cell(50, 1, f"TBD - Security Approver\nSignature Over Printed Name / Date")
            st.info(f"🔒 LUZ Region detected - Security Approver set to: TBD (Please update when known)")
        
        # --- APPLY CALIBRI 6 FONT TO ALL CELLS ---
        calibri_6 = Font(name="Calibri", size=6)
        calibri_6_underline = Font(name="Calibri", size=6, underline="single")
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
            for cell in row:
                if cell.value:
                    if cell.row >= 48 and cell.row <= 55:
                        safe_set_font(cell.row, cell.column, calibri_6_underline)
                    else:
                        safe_set_font(cell.row, cell.column, calibri_6)
        
        safe_set_font(48, 1, calibri_6_underline)
        safe_set_font(50, 1, calibri_6_underline)
        safe_set_font(41, 1, calibri_6)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        wb.close()
        gc.collect()
        
        return buffer
        
    except Exception as e:
        st.error(f"Error creating RAAWA file: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

# --- SITE GROUPING FUNCTIONS ---
def split_sites_by_territory_and_fm(matching_sites):
    """Split sites by unique combinations of Territory and Facility Manager, then further split if > 10 sites per group"""
    unique_combos = get_unique_combinations(matching_sites)
    
    final_groups = []
    
    for key, combo in unique_combos.items():
        sites_list = combo['sites']
        territory = combo['territory']
        facility_manager = combo['facility_manager']
        
        sites_df = pd.DataFrame(sites_list)
        
        if len(sites_df) > 10:
            num_batches = math.ceil(len(sites_df) / 10)
            for batch_num in range(num_batches):
                start_idx = batch_num * 10
                end_idx = min((batch_num + 1) * 10, len(sites_df))
                batch_sites = sites_df.iloc[start_idx:end_idx]
                
                final_groups.append({
                    'dataframe': batch_sites,
                    'territory': territory,
                    'facility_manager': facility_manager,
                    'batch_num': batch_num + 1,
                    'total_batches': num_batches,
                    'original_count': len(sites_df)
                })
        else:
            final_groups.append({
                'dataframe': sites_df,
                'territory': territory,
                'facility_manager': facility_manager,
                'batch_num': 1,
                'total_batches': 1,
                'original_count': len(sites_df)
            })
    
    return final_groups

def get_unique_combinations(matching_sites):
    """Get unique combinations of Territory and NEW ENGINEER_AH that must be separated"""
    combinations = {}
    
    for _, row in matching_sites.iterrows():
        territory = str(row.get("TERRITORY", "")).strip()
        facility_manager = str(row.get("NEW ENGINEER_AH", "")).strip()
        
        if facility_manager in ['N/A', 'nan', '']:
            facility_manager = "Unassigned FM"
        
        key = f"{territory}_{facility_manager}"
        
        if key not in combinations:
            combinations[key] = {
                'territory': territory,
                'facility_manager': facility_manager,
                'sites': []
            }
        
        combinations[key]['sites'].append(row)
    
    return combinations

def check_conflicts(matching_sites):
    """Check for conflicts in territories and facility managers"""
    unique_combos = get_unique_combinations(matching_sites)
    territories = set([combo['territory'] for combo in unique_combos.values()])
    facility_managers = set([combo['facility_manager'] for combo in unique_combos.values()])
    
    territory_conflict = '7' in territories and '8' in territories
    fm_conflict = len(facility_managers) > 1
    
    return {
        'territory_conflict': territory_conflict,
        'fm_conflict': fm_conflict,
        'unique_territories': territories,
        'unique_facility_managers': facility_managers,
        'num_combinations': len(unique_combos)
    }

# --- REQUISITIONER FUNCTION (AUTO-DETECT) ---
def get_requisitioner_for_territory_and_project(territory, project, region):
    """Get requisitioner profile based on territory, project, and region (auto mode)"""
    if df_req_db is not None and 'Territory no.' in df_req_db.columns:
        # Try exact match first
        matching_reqs = df_req_db[
            (df_req_db['Territory no.'] == territory) &
            (df_req_db['Project'] == project) &
            (df_req_db['Region'] == region)
        ]
        
        if not matching_reqs.empty:
            req_row = matching_reqs.iloc[0]
            return {
                "name": str(req_row.get("Name", "N/A")),
                "dept": str(req_row.get("Dept./Group", "N/A")),
                "id": format_id_number(req_row.get("ID #", "N/A")),
                "contact": format_contact_number(req_row.get("Contact No.", "N/A"))
            }
        
        # If exact match fails, try partial match
        if project:
            base_project = project.split(' - ')[0] if ' - ' in project else project
            
            matching_reqs = df_req_db[
                (df_req_db['Territory no.'] == territory) &
                (df_req_db['Project'].str.contains(base_project, case=False, na=False)) &
                (df_req_db['Region'] == region)
            ]
            
            if not matching_reqs.empty:
                req_row = matching_reqs.iloc[0]
                return {
                    "name": str(req_row.get("Name", "N/A")),
                    "dept": str(req_row.get("Dept./Group", "N/A")),
                    "id": format_id_number(req_row.get("ID #", "N/A")),
                    "contact": format_contact_number(req_row.get("Contact No.", "N/A"))
                }
    
    # Fallback - try without project filter
    if df_req_db is not None and 'Territory no.' in df_req_db.columns:
        matching_reqs = df_req_db[
            (df_req_db['Territory no.'] == territory) &
            (df_req_db['Region'] == region)
        ]
        
        if not matching_reqs.empty:
            req_row = matching_reqs.iloc[0]
            return {
                "name": str(req_row.get("Name", "N/A")),
                "dept": str(req_row.get("Dept./Group", "N/A")),
                "id": format_id_number(req_row.get("ID #", "N/A")),
                "contact": format_contact_number(req_row.get("Contact No.", "N/A"))
            }
    
    return {
        "name": f"Territory {territory} Engineer",
        "dept": f"TERRITORY {territory}",
        "id": "N/A",
        "contact": "N/A"
    }

# --- REQUISITIONER SELECTION FUNCTION (MANUAL/DATABASE) ---
def get_requisitioner_selection(req_db):
    """Display requisitioner selection interface with manual and database options"""
    
    # Initialize session state for requisitioner
    if 'selected_requisitioner' not in st.session_state:
        st.session_state.selected_requisitioner = {
            "name": "N/A",
            "dept": "N/A",
            "id": "N/A",
            "contact": "N/A"
        }
    
    if 'req_selection_method' not in st.session_state:
        st.session_state.req_selection_method = "Auto (Territory-based)"
    
    # Requisitioner selection method
    req_method = st.radio(
        "Requisitioner Selection Method:",
        ["Auto (Territory-based)", "Manual Input", "Load from Database", "Mixed (Manual + Database)"],
        horizontal=True,
        key="req_method_main"
    )
    
    st.session_state.req_selection_method = req_method
    
    # Store manual input values in session state
    for key in ['manual_req_name', 'manual_req_dept', 'manual_req_id', 'manual_req_contact']:
        if key not in st.session_state:
            st.session_state[key] = ""
    
    if req_method == "Manual Input":
        col1, col2 = st.columns(2)
        with col1:
            manual_name = st.text_input("Requisitioner Name:", value=st.session_state.manual_req_name, key="manual_req_name_input")
            manual_dept = st.text_input("Department/Group:", value=st.session_state.manual_req_dept, key="manual_req_dept_input")
        with col2:
            manual_id = st.text_input("ID Number:", value=st.session_state.manual_req_id, key="manual_req_id_input")
            manual_contact = st.text_input("Contact Number:", value=st.session_state.manual_req_contact, key="manual_req_contact_input")
        
        # Update session state with current values
        st.session_state.manual_req_name = manual_name
        st.session_state.manual_req_dept = manual_dept
        st.session_state.manual_req_id = manual_id
        st.session_state.manual_req_contact = manual_contact
        
        if st.button("✅ Set Requisitioner", key="set_manual_req"):
            if manual_name:
                st.session_state.selected_requisitioner = {
                    "name": manual_name,
                    "dept": manual_dept if manual_dept else "N/A",
                    "id": format_id_number(manual_id) if manual_id else "N/A",
                    "contact": format_contact_number(manual_contact) if manual_contact else "N/A"
                }
                st.success(f"✅ Requisitioner set to: {manual_name}")
        
        # Show current manual requisitioner
        if st.session_state.selected_requisitioner["name"] != "N/A":
            st.info(f"📋 Current Requisitioner: **{st.session_state.selected_requisitioner['name']}**")
    
    elif req_method == "Load from Database":
        if req_db is not None and not req_db.empty:
            # Create searchable dropdown
            st.subheader("Select from Requisitioner Database")
            
            # Create a formatted list for display
            req_options = []
            req_dict = {}
            
            for idx, row in req_db.iterrows():
                name = str(row.get('Name', 'N/A'))
                dept = str(row.get('Dept./Group', 'N/A'))
                id_num = str(row.get('ID #', 'N/A'))
                contact = str(row.get('Contact No.', 'N/A'))
                territory = str(row.get('Territory no.', 'N/A'))
                region = str(row.get('Region', 'N/A'))
                project = str(row.get('Project', 'N/A'))
                
                display_text = f"{name} - {dept} (Territory: {territory}, Region: {region})"
                req_options.append(display_text)
                req_dict[display_text] = {
                    "name": name,
                    "dept": dept,
                    "id": id_num,
                    "contact": contact,
                    "territory": territory,
                    "region": region,
                    "project": project
                }
            
            # Add search/filter
            search_term = st.text_input("🔍 Search Requisitioner:", key="req_search")
            
            filtered_options = req_options
            if search_term:
                filtered_options = [opt for opt in req_options if search_term.lower() in opt.lower()]
            
            if filtered_options:
                selected_option = st.selectbox(
                    "Select Requisitioner:",
                    options=filtered_options,
                    key="req_select"
                )
                
                if st.button("✅ Select Requisitioner", key="set_db_req"):
                    if selected_option in req_dict:
                        req_data = req_dict[selected_option]
                        st.session_state.selected_requisitioner = {
                            "name": req_data["name"],
                            "dept": req_data["dept"],
                            "id": format_id_number(req_data["id"]),
                            "contact": format_contact_number(req_data["contact"])
                        }
                        st.success(f"✅ Requisitioner set to: {req_data['name']}")
            else:
                st.warning("No requisitioners match your search")
        else:
            st.warning("No requisitioner database loaded. Please check Requisitioner.xlsx")
        
        # Show current requisitioner
        if st.session_state.selected_requisitioner["name"] != "N/A":
            st.info(f"📋 Current Requisitioner: **{st.session_state.selected_requisitioner['name']}**")
    
    elif req_method == "Mixed (Manual + Database)":
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📁 Database Selection")
            if req_db is not None and not req_db.empty:
                # Create a formatted list for display
                req_options = []
                req_dict = {}
                
                for idx, row in req_db.iterrows():
                    name = str(row.get('Name', 'N/A'))
                    dept = str(row.get('Dept./Group', 'N/A'))
                    id_num = str(row.get('ID #', 'N/A'))
                    contact = str(row.get('Contact No.', 'N/A'))
                    territory = str(row.get('Territory no.', 'N/A'))
                    region = str(row.get('Region', 'N/A'))
                    project = str(row.get('Project', 'N/A'))
                    
                    display_text = f"{name} - {dept} (Territory: {territory}, Region: {region})"
                    req_options.append(display_text)
                    req_dict[display_text] = {
                        "name": name,
                        "dept": dept,
                        "id": id_num,
                        "contact": contact,
                        "territory": territory,
                        "region": region,
                        "project": project
                    }
                
                selected_option = st.selectbox(
                    "Select Requisitioner:",
                    options=req_options,
                    key="req_select_mixed"
                )
                
                if st.button("✅ Select from Database", key="set_db_req_mixed"):
                    if selected_option in req_dict:
                        req_data = req_dict[selected_option]
                        st.session_state.selected_requisitioner = {
                            "name": req_data["name"],
                            "dept": req_data["dept"],
                            "id": format_id_number(req_data["id"]),
                            "contact": format_contact_number(req_data["contact"])
                        }
                        st.success(f"✅ Requisitioner set to: {req_data['name']}")
            else:
                st.warning("No requisitioner database loaded")
        
        with col2:
            st.subheader("✏️ Manual Entry")
            manual_name = st.text_input("Requisitioner Name:", value=st.session_state.manual_req_name, key="manual_req_name_mixed")
            manual_dept = st.text_input("Department/Group:", value=st.session_state.manual_req_dept, key="manual_req_dept_mixed")
            manual_id = st.text_input("ID Number:", value=st.session_state.manual_req_id, key="manual_req_id_mixed")
            manual_contact = st.text_input("Contact Number:", value=st.session_state.manual_req_contact, key="manual_req_contact_mixed")
            
            # Update session state with current values
            st.session_state.manual_req_name = manual_name
            st.session_state.manual_req_dept = manual_dept
            st.session_state.manual_req_id = manual_id
            st.session_state.manual_req_contact = manual_contact
            
            if st.button("✅ Set Manual Requisitioner", key="set_manual_req_mixed"):
                if manual_name:
                    st.session_state.selected_requisitioner = {
                        "name": manual_name,
                        "dept": manual_dept if manual_dept else "N/A",
                        "id": format_id_number(manual_id) if manual_id else "N/A",
                        "contact": format_contact_number(manual_contact) if manual_contact else "N/A"
                    }
                    st.success(f"✅ Requisitioner set to: {manual_name}")
        
        # Show current requisitioner
        if st.session_state.selected_requisitioner["name"] != "N/A":
            st.info(f"📋 Current Requisitioner: **{st.session_state.selected_requisitioner['name']}**")
    
    else:  # Auto (Territory-based)
        st.info("ℹ️ Requisitioner will be automatically determined based on Territory, Project, and Region")
        if st.session_state.selected_requisitioner["name"] != "N/A":
            st.info(f"📋 Current Requisitioner: **{st.session_state.selected_requisitioner['name']}**")
    
    return st.session_state.selected_requisitioner

# --- MAIN APP ---
# Check files first
if not check_required_files():
    st.stop()

# Load databases
df_db, df_req_db, df_engr_tech_db = load_databases()

# --- SIDEBAR NAVIGATION ---
st.sidebar.markdown("""
<div class="sidebar-content">
    <h2 style="text-align: center; color: #667eea;">📋 RAAWA Generator</h2>
    <hr>
</div>
""", unsafe_allow_html=True)

page = st.sidebar.radio(
    "Navigate",
    ["🏠 Main Generator", "ℹ️ About & Developer"],
    format_func=lambda x: x
)

st.sidebar.markdown("---")
st.sidebar.markdown("""
<div style="text-align: center; padding: 1rem;">
    <small>Version 3.0.0</small><br>
    <small>© 2026 RAAWA Generator</small><br>
    <small>✨ Multi-Region Support (LUZ, VIS, MIN)</small>
</div>
""", unsafe_allow_html=True)

# --- ABOUT PAGE ---
if page == "ℹ️ About & Developer":
    st.markdown("""
    <div class="main-header">
        <h1>📄 About RAAWA Generator</h1>
        <p>Professional Multi-Site RAAWA Document Automation System</p>
        <p><strong>🌍 Now with LUZ, VIS, and MIN Database Support!</strong></p>
    </div>
    """, unsafe_allow_html=True)
    
    # Developer and Mission section with dark background
    st.markdown("""
    <div style="background: linear-gradient(135deg, #2d3436 0%, #1a1a2e 100%); 
                padding: 2rem; 
                border-radius: 15px; 
                margin: 1rem 0;
                color: white;">
        <div style="display: flex; gap: 2rem; flex-wrap: wrap;">
            <div style="flex: 1; min-width: 250px;">
                <h2 style="color: #667eea;">👨‍💻 Developer</h2>
                <hr style="border-color: #667eea;">
                <h3 style="color: #a29bfe;">John Carlo Rabanes</h3>
                <p><strong style="color: #fd79a8;">OLT Rollout Engineer</strong></p>
                <p>📞 09669343065</p>
                <p>📧 rabanes.johncarlo4@gmail.com</p>
                <p>🏢 Nokia Shanghai Bell</p>
            </div>
            <div style="flex: 1.5; min-width: 300px;">
                <h2 style="color: #667eea;">🎯 Mission</h2>
                <hr style="border-color: #667eea;">
                <p style="font-size: 1.1rem;">To streamline and automate the RAAWA (Request for Authority to Access Work Area) document generation process, reducing manual effort and eliminating errors in multi-site telecommunications infrastructure projects.</p>
                <br>
                <h2 style="color: #667eea;">💡 Vision</h2>
                <hr style="border-color: #667eea;">
                <p style="font-size: 1.1rem;">To become the standard tool for telecommunications field operations, enabling engineers to generate compliant documentation in minutes instead of hours.</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Key Advantages section
    st.markdown("## 🌟 Key Advantages")
    col1, col2, col3 = st.columns(3)
    
    advantages = [
        ("⏱️ Time Saving", "Generate complete RAAWA forms in seconds instead of 30+ minutes"),
        ("🎯 Zero Errors", "Eliminate manual data entry mistakes and formatting issues"),
        ("📊 Unlimited Sites", "Handle any number of sites with automatic batching")
    ]
    
    for idx, (title, desc) in enumerate(advantages):
        with [col1, col2, col3][idx]:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        padding: 1.5rem;
                        border-radius: 15px;
                        margin: 0.5rem 0;
                        color: white;
                        min-height: 150px;
                        display: flex;
                        flex-direction: column;
                        justify-content: center;">
                <h3 style="margin: 0; font-size: 1.3rem;">{title}</h3>
                <p style="margin: 0.5rem 0 0 0; font-size: 0.95rem; opacity: 0.9;">{desc}</p>
            </div>
            """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    advantages2 = [
        ("🔄 Auto-Routing", "Automatic requisitioner mapping based on territory"),
        ("📁 Multi-Region DB", "Load sites from LUZ, VIS, and MIN databases"),
        ("🔒 Smart Batching", "Auto-splits into multiple RAAWAs when exceeding 10 sites")
    ]
    
    for idx, (title, desc) in enumerate(advantages2):
        with [col1, col2, col3][idx]:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        padding: 1.5rem;
                        border-radius: 15px;
                        margin: 0.5rem 0;
                        color: white;
                        min-height: 150px;
                        display: flex;
                        flex-direction: column;
                        justify-content: center;">
                <h3 style="margin: 0; font-size: 1.3rem;">{title}</h3>
                <p style="margin: 0.5rem 0 0 0; font-size: 0.95rem; opacity: 0.9;">{desc}</p>
            </div>
            """, unsafe_allow_html=True)
    
    # Features section with cards
    st.markdown("## ⚡ Features")
    
    features = [
        ("🏗️ Unlimited Sites", "Select any number of sites from LUZ, VIS, or MIN databases", "✅"),
        ("👥 Personnel Management", "Manual, database, or mixed input modes", "✅"),
        ("📝 Requisitioner Selection", "Manual, database, or auto territory-based selection", "✅"),
        ("📝 Project & Region Aware", "Filter requisitioners and teams by project/region (LUZ, VIS, MIN)", "✅"),
        ("🏢 Company Filtering", "Select personnel by company with region awareness", "✅"),
        ("📄 Professional Output", "Perfectly formatted Excel with auto font sizing", "✅"),
        ("🔍 Search & Filter", "Quick personnel and requisitioner search from database", "✅"),
        ("💾 Batch Download", "Download multiple RAAWA files at once", "✅"),
        ("🎨 Clean Interface", "User-friendly with professional design", "✅"),
        ("📊 Real-time Preview", "See selected sites and personnel before generation", "✅"),
        ("🔄 Session Management", "Persistent data across navigation", "✅"),
        ("📑 Auto-Batching", "Automatically splits into 10-site batches per signatory group", "✅"),
        ("📅 Flexible Dates", "Choose start date manually or use current date", "✅"),
        ("📞 Auto-Format Contact", "Automatically formats contact numbers to start with 0", "✅"),
        ("🆔 Clean ID Numbers", "Removes decimal places from ID numbers", "✅"),
        ("🌍 Multi-Region", "Support for LUZ, VIS, and MIN site databases", "✅")
    ]
    
    # Display features in a grid
    for i in range(0, len(features), 2):
        cols = st.columns(2)
        for j in range(2):
            if i + j < len(features):
                feature = features[i + j]
                with cols[j]:
                    st.markdown(f"""
                    <div style="background: white;
                                padding: 1rem;
                                border-radius: 10px;
                                margin: 0.5rem 0;
                                border-left: 4px solid #667eea;
                                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
                                min-height: 80px;
                                display: flex;
                                align-items: center;">
                        <div>
                            <h3 style="margin: 0; color: #2d3436; font-size: 1.05rem;">
                                {feature[0]} <span style="color: #00b894;">{feature[2]}</span>
                            </h3>
                            <p style="margin: 0.2rem 0 0 0; color: #636e72; font-size: 0.9rem;">{feature[1]}</p>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
    
    # Technology Stack with dark cards
    st.markdown("## 🛠️ Technology Stack")
    tech_col1, tech_col2, tech_col3, tech_col4 = st.columns(4)
    
    tech_stack = [
        ("🐍 Python", "Core Logic"),
        ("🎈 Streamlit", "Web Framework"),
        ("📊 Pandas", "Data Processing"),
        ("📝 OpenPyXL", "Excel Generation")
    ]
    
    for idx, (icon, name) in enumerate(tech_stack):
        with [tech_col1, tech_col2, tech_col3, tech_col4][idx]:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #2d3436 0%, #1a1a2e 100%);
                        padding: 1.5rem;
                        border-radius: 15px;
                        margin: 0.5rem 0;
                        color: white;
                        text-align: center;
                        min-height: 120px;
                        display: flex;
                        flex-direction: column;
                        justify-content: center;">
                <h2 style="margin: 0; font-size: 2.5rem;">{icon}</h2>
                <h3 style="margin: 0.5rem 0 0 0; font-size: 1rem; color: #a29bfe;">{name}</h3>
            </div>
            """, unsafe_allow_html=True)
    
    # Support & Contact with gradient background
    st.markdown("## 📞 Support & Contact")
    st.markdown("""
    <div style="background: linear-gradient(135deg, #dfe6e9 0%, #b2bec3 100%);
                padding: 2rem;
                border-radius: 15px;
                margin: 1rem 0;
                border-left: 4px solid #667eea;">
        <h3 style="color: #2d3436;">Need Help?</h3>
        <p>📧 Email: <a href="mailto:rabanes.johncarlo4@gmail.com" style="color: #0984e3; text-decoration: none;">rabanes.johncarlo4@gmail.com</a></p>
        <p>📱 Phone: 09669343065</p>
        <p>🏢 Company: Nokia Shanghai Bell</p>
        <hr style="border-color: #636e72;">
        <p style="font-size: 0.9rem; color: #2d3436;"><small>For technical support, feature requests, or bug reports, please reach out via email.</small></p>
    </div>
    """, unsafe_allow_html=True)

# --- MAIN GENERATOR PAGE ---
else:
    if df_db is not None and df_req_db is not None:
        st.markdown("""
        <div class="main-header">
            <h1>📄 Automated Multi-Site RAAWA Generator</h1>
            <p>Select any number of sites, populate the personnel manifest, and generate perfectly styled RAAWA forms instantly.<br>
            <strong>🌍 Multi-Region (LUZ, VIS, MIN) - Auto-batches into groups of 10 per signatory</strong></p>
        </div>
        """, unsafe_allow_html=True)
        
        # --- Display Database Statistics ---
        with st.expander("📊 Database Statistics", expanded=False):
            col1, col2, col3 = st.columns(3)
            with col1:
                min_count = len(df_db[df_db['REGION'] == 'MIN']) if 'REGION' in df_db.columns else 0
                st.metric("🗺️ MIN Sites", min_count)
            with col2:
                vis_count = len(df_db[df_db['REGION'] == 'VIS']) if 'REGION' in df_db.columns else 0
                st.metric("🗺️ VIS Sites", vis_count)
            with col3:
                luz_count = len(df_db[df_db['REGION'] == 'LUZ']) if 'REGION' in df_db.columns else 0
                st.metric("🗺️ LUZ Sites", luz_count)
            st.caption(f"Total Sites: {len(df_db)}")
        
        # --- STEP 1: PROJECT & REGION SELECTION ---
        st.markdown("## 📋 Step 1: Project & Region Configuration")
        
        project_list = []
        
        if df_req_db is not None and 'Project' in df_req_db.columns:
            project_list = sorted(df_req_db['Project'].dropna().unique())
            project_list = [p for p in project_list if str(p).strip() != '' and str(p).strip() != 'nan']
        else:
            project_list = ['Default']
        
        col_project, col_region = st.columns(2)
        
        with col_project:
            selected_project = st.selectbox(
                "Select Project:",
                options=project_list,
                help="Choose the project for the RAAWA"
            )
        
        with col_region:
            # HARDCODE REGIONS - Always show LUZ, VIS, MIN regardless of database
            site_regions = ['LUZ', 'VIS', 'MIN']
            
            selected_region = st.selectbox(
                "Select Region:",
                options=['All Regions'] + site_regions,
                help="Filter sites by region (LUZ, VIS, MIN)"
            )
        
        st.markdown("---")
        
        # --- STEP 1.5: REQUISITIONER SELECTION ---
        st.markdown("## 👤 Step 1.5: Requisitioner Selection")
        st.markdown("Select the requisitioner for this RAAWA (can be manual, from database, or auto)")
        
        selected_req_profile = get_requisitioner_selection(df_req_db)
        
        st.markdown("---")
        
        # --- STEP 2: MULTI-SITE SELECTION INPUT ---
        st.markdown("## 📍 Step 2: Site Selection")
        st.markdown("Select the sites for your RAAWA request (unlimited number supported)")
        
        # Filter by selected region - handle case where region might not exist in db yet
        if selected_region != 'All Regions' and df_db is not None and 'REGION' in df_db.columns:
            filtered_db = df_db[df_db['REGION'].str.upper() == selected_region.upper()]
        else:
            filtered_db = df_db
        
        # Check if filtered_db is empty and show a helpful message
        if filtered_db.empty and selected_region != 'All Regions':
            st.warning(f"⚠️ No sites found for region: **{selected_region}**. Please check if the database file for this region is loaded correctly.")
        
        # Cache the plaid list for performance
        plaid_list = sorted(filtered_db['PLAID'].unique()) if not filtered_db.empty else []
        
        if plaid_list:
            selected_plaids = st.multiselect(
                "Search or Select Site IDs (PLAID):", 
                options=plaid_list,
                default=[]
            )
                
            matching_sites = filtered_db[filtered_db['PLAID'].isin(selected_plaids)]
        else:
            matching_sites = pd.DataFrame()
            st.info("No sites available for the selected region.")
        
        if not matching_sites.empty:
            conflicts = check_conflicts(matching_sites)
            
            if conflicts['territory_conflict']:
                st.warning("⚠️ **Territory Limitation Detected:** You have selected sites from both Territory 7 and Territory 8. These cannot be combined in a single RAAWA.")
            
            if conflicts['fm_conflict']:
                st.warning(f"⚠️ **Facility Manager Conflict Detected:** Multiple Facility Managers found ({', '.join(conflicts['unique_facility_managers'])}). Each unique Territory + Facility Manager combination will get its own RAAWA.")
            
            st.markdown("### 📋 Selected Sites Preview")
            
            preview_cols = ['PLAID', 'SITE', 'REGION', 'TERRITORY', 'NEW ENGINEER_AH', 'CONTACT NUMBER', 'SITE_ADD']
            available_cols = [col for col in preview_cols if col in matching_sites.columns]
            st.dataframe(
                matching_sites[available_cols],
                hide_index=True,
                use_container_width=True
            )
            
            unique_combos = get_unique_combinations(matching_sites)
            
            total_files = 0
            batch_details = []
            for combo in unique_combos.values():
                num_sites = len(combo['sites'])
                num_files = math.ceil(num_sites / 10)
                total_files += num_files
                batch_details.append({
                    'territory': combo['territory'],
                    'fm': combo['facility_manager'],
                    'sites': num_sites,
                    'files': num_files
                })
            
            st.info(f"📋 **Generation Plan:** The system will create **{total_files} separate RAAWA files**")
            
            with st.expander("📊 View Detailed RAAWA Breakdown"):
                for detail in batch_details:
                    if detail['files'] > 1:
                        st.markdown(f"""
                        **Territory {detail['territory']} - {detail['fm']}**:
                        - {detail['sites']} sites → {detail['files']} RAAWA files (batches of 10)
                        """)
                    else:
                        st.markdown(f"""
                        **Territory {detail['territory']} - {detail['fm']}**:
                        - {detail['sites']} sites → {detail['files']} RAAWA file
                        """)
                    st.markdown("---")
                
        else:
            if plaid_list:
                st.info("💡 Please select at least one site to begin")
                st.stop()
            else:
                st.warning("No sites available for the selected region. Please check your database files.")
                st.stop()

        st.markdown("---")

        # --- STEP 3: ACCESS SCOPE & DEPLOYMENTS ---
        st.markdown("## 📝 Step 3: Access Scope & Deployment Details")
        
        scope_col, date_col = st.columns(2)
        with scope_col:
            scope_of_work = st.text_area(
                "Nature of Access / Detailed Scope of Work:",
                value="SITE SURVEY, INSTALLATION, INTEGRATION, AND ACCEPTANCE TESTING OF NOKIA OLT.",
                height=100
            )
        with date_col:
            date_option = st.radio(
                "Date Selection Method:",
                ["Use Current Date", "Manual Date Entry"],
                horizontal=True
            )
            
            if date_option == "Use Current Date":
                start_date = datetime.now()
                st.info(f"📅 Start Date: **{start_date.strftime('%Y-%m-%d')}** (Current Date)")
            else:
                start_date = st.date_input(
                    "Select Start Date:",
                    value=datetime.now().date(),
                    min_value=datetime.now().date(),
                    help="Choose the effective start date for the authorization"
                )
                start_date = datetime.combine(start_date, datetime.min.time())
                st.caption(f"Selected Start Date: **{start_date.strftime('%Y-%m-%d')}**")
            
            validity_days = st.number_input(
                "Authorization Validity (Days):", 
                min_value=1, 
                max_value=365, 
                value=30,
                help="Number of days the authorization will be valid from the start date"
            )
            
            end_date = start_date + timedelta(days=int(validity_days))
            
            st.success(f"📅 Form window coverage: **{start_date.strftime('%Y-%m-%d')}** to **{end_date.strftime('%Y-%m-%d')}**")
            st.caption(f"Total validity: {validity_days} days")

        st.markdown("## 👥 Step 4: Team Manifest")
        st.markdown("Add the engineers and technicians who will perform the work")
        
        # --- PERSONNEL SELECTION METHOD ---
        col_method, col_empty = st.columns([2, 3])
        with col_method:
            selection_method = st.radio(
                "Personnel Selection Method:",
                ["Manual Input", "Load from Database", "Mixed (Database + Manual)"],
                horizontal=True,
                key="selection_method_main"
            )
        
        if 'personnel_list' not in st.session_state:
            st.session_state.personnel_list = []
        
        # Store selected indices in session state to persist across reruns
        if 'selected_indices_temp' not in st.session_state:
            st.session_state.selected_indices_temp = []
        
        if selection_method == "Load from Database":
            if df_engr_tech_db is None or df_engr_tech_db.empty:
                st.warning("⚠️ No personnel records found in EngrTech.xlsx")
                st.info("""
                **Expected format for EngrTech.xlsx:**
                | Name | Company | SEC ID | REGION |
                |------|---------|--------|--------|
                | John Doe | Nokia | 12345 | LUZ |
                | Jane Smith | Huawei | 67890 | VIS |
                """)
            else:
                # Display the full count of personnel loaded
                st.success(f"📋 Found {len(df_engr_tech_db)} personnel records in database")
                
                # Show region breakdown
                if 'Region' in df_engr_tech_db.columns:
                    region_counts = df_engr_tech_db['Region'].value_counts().to_dict()
                    region_summary = ", ".join([f"{k}: {v}" for k, v in region_counts.items()])
                    st.info(f"📊 Region breakdown: {region_summary}")
                
                with st.expander("📁 Select Personnel from Database", expanded=True):
                    # Get unique companies and regions
                    company_list = sorted(df_engr_tech_db['Company'].unique())
                    company_list = [c for c in company_list if str(c).strip() != '']
                    
                    region_list = []
                    if 'Region' in df_engr_tech_db.columns:
                        region_list = sorted(df_engr_tech_db['Region'].unique())
                        region_list = [r for r in region_list if str(r).strip() != '']
                        # Move 'N/A' to the end if it exists
                        if 'N/A' in region_list:
                            region_list.remove('N/A')
                            region_list.append('N/A')
                        region_list = ['All Regions'] + region_list
                    else:
                        region_list = ['All Regions']
                    
                    # Filters in columns
                    col_filters = st.columns(3)
                    
                    with col_filters[0]:
                        if company_list:
                            selected_companies = st.multiselect(
                                "Filter by Company:",
                                options=['All Companies'] + company_list,
                                default=['All Companies'],
                                key="company_filter_main"
                            )
                        else:
                            selected_companies = ['All Companies']
                            st.info("No companies found")
                    
                    with col_filters[1]:
                        if region_list:
                            selected_regions = st.multiselect(
                                "Filter by Region:",
                                options=region_list,
                                default=['All Regions'],
                                key="region_filter_main"
                            )
                        else:
                            selected_regions = ['All Regions']
                            st.info("No regions found")
                    
                    with col_filters[2]:
                        search_term = st.text_input("🔍 Search:", key="personnel_search_main")
                    
                    # Apply filters - FIXED LOGIC
                    filtered_engr = df_engr_tech_db.copy()
                    
                    # Company filter - FIXED: Only filter if not "All Companies"
                    if selected_companies and 'All Companies' not in selected_companies:
                        filtered_engr = filtered_engr[filtered_engr['Company'].isin(selected_companies)]
                    
                    # Region filter - FIXED: Only filter if not "All Regions"
                    if selected_regions and 'All Regions' not in selected_regions:
                        if 'Region' in filtered_engr.columns:
                            filtered_engr = filtered_engr[filtered_engr['Region'].isin(selected_regions)]
                    
                    # Search filter
                    if search_term:
                        filtered_engr = filtered_engr[
                            filtered_engr['Name'].astype(str).str.contains(search_term, case=False, na=False) |
                            filtered_engr['Company'].astype(str).str.contains(search_term, case=False, na=False) |
                            filtered_engr['ID No'].astype(str).str.contains(search_term, case=False, na=False)
                        ]
                    
                    # Show count
                    st.info(f"Showing {len(filtered_engr)} of {len(df_engr_tech_db)} personnel")
                    
                    # Display filtered data
                    display_cols = []
                    if 'Name' in filtered_engr.columns:
                        display_cols.append('Name')
                    if 'Company' in filtered_engr.columns:
                        display_cols.append('Company')
                    if 'ID No' in filtered_engr.columns:
                        display_cols.append('ID No')
                    if 'Region' in filtered_engr.columns:
                        display_cols.append('Region')
                    
                    if not filtered_engr.empty and display_cols:
                        st.dataframe(
                            filtered_engr[display_cols],
                            hide_index=True,
                            use_container_width=True
                        )
                        
                        # Get the list of indices
                        options_list = filtered_engr.index.tolist()
                        
                        if options_list:
                            # Create a safe format function
                            def format_person(row_idx):
                                try:
                                    name = filtered_engr.loc[row_idx, 'Name']
                                    company = filtered_engr.loc[row_idx, 'Company']
                                    id_no = filtered_engr.loc[row_idx, 'ID No']
                                    return f"{name} - {company} (ID: {id_no})"
                                except:
                                    return f"Person {row_idx}"
                            
                            # Multi-select with filtered data
                            selected_indices = st.multiselect(
                                "Select Engineers/Technicians to add:",
                                options=options_list,
                                format_func=format_person,
                                key="selected_personnel_db_main"
                            )
                            
                            # Store selected indices in session state
                            st.session_state.selected_indices_temp = selected_indices
                            
                            col_add, col_add_all, col_clear = st.columns([1, 1, 2])
                            with col_add:
                                if st.button("➕ Add Selected", use_container_width=True):
                                    if selected_indices:
                                        added_count = 0
                                        for idx in selected_indices:
                                            try:
                                                person = filtered_engr.loc[idx]
                                                if not any(p['name'] == person['Name'] for p in st.session_state.personnel_list):
                                                    st.session_state.personnel_list.append({
                                                        "name": person['Name'],
                                                        "company": person['Company'],
                                                        "id_no": person['ID No']
                                                    })
                                                    added_count += 1
                                            except:
                                                pass
                                        if added_count > 0:
                                            st.success(f"✅ Added {added_count} personnel to manifest!")
                                        else:
                                            st.info("All selected personnel are already in the manifest.")
                                    else:
                                        st.warning("Please select personnel first")
                            
                            with col_add_all:
                                if st.button("➕ Add All Visible", use_container_width=True):
                                    # Get all indices from the filtered list
                                    all_indices = filtered_engr.index.tolist()
                                    if all_indices:
                                        added_count = 0
                                        for idx in all_indices:
                                            try:
                                                person = filtered_engr.loc[idx]
                                                if not any(p['name'] == person['Name'] for p in st.session_state.personnel_list):
                                                    st.session_state.personnel_list.append({
                                                        "name": person['Name'],
                                                        "company": person['Company'],
                                                        "id_no": person['ID No']
                                                    })
                                                    added_count += 1
                                            except:
                                                pass
                                        if added_count > 0:
                                            st.success(f"✅ Added all {added_count} visible personnel to manifest!")
                                            # Clear the selection after adding all
                                            st.session_state.selected_indices_temp = []
                                        else:
                                            st.info("All visible personnel are already in the manifest.")
                                    else:
                                        st.warning("No personnel to add.")
                        else:
                            st.warning("No personnel match the current filters")
                    else:
                        st.warning("No personnel match the current filters")
                
                # Manual entry section
                st.markdown("---")
                st.subheader("Or Add Manually")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    manual_name = st.text_input("Name:", key="manual_name_db")
                with col2:
                    manual_company = st.text_input("Company:", key="manual_company_db")
                with col3:
                    manual_id = st.text_input("ID No.:", key="manual_id_db")
                
                if st.button("➕ Add Manual Entry", key="add_manual_db"):
                    if manual_name:
                        if not any(p['name'] == manual_name for p in st.session_state.personnel_list):
                            st.session_state.personnel_list.append({
                                "name": manual_name,
                                "company": manual_company,
                                "id_no": manual_id
                            })
                            st.success(f"✅ Added {manual_name} to manifest!")
                        else:
                            st.warning(f"{manual_name} is already in the manifest.")
        
        if selection_method == "Manual Input":
            # Track visibility dynamically up to 40 entries
            for i in range(1, 41):
                if f"p_show_{i}" not in st.session_state:
                    st.session_state[f"p_show_{i}"] = True if i <= 4 else False
                    
            visible_count = sum(1 for i in range(1, 41) if st.session_state[f"p_show_{i}"])
            
            # Add "Add All" button at the top
            col_add_all_manual, col_spacer = st.columns([1, 3])
            with col_add_all_manual:
                if st.button("➕ Add All Visible Fields", key="add_all_manual", use_container_width=True):
                    added_count = 0
                    for i in range(1, 41):
                        if st.session_state[f"p_show_{i}"]:
                            name_key = f"name_{i}"
                            comp_key = f"comp_{i}"
                            id_key = f"id_{i}"
                            
                            if name_key in st.session_state and st.session_state[name_key]:
                                name = st.session_state[name_key]
                                if not any(p['name'] == name for p in st.session_state.personnel_list):
                                    st.session_state.personnel_list.append({
                                        "name": name,
                                        "company": st.session_state.get(comp_key, ""),
                                        "id_no": st.session_state.get(id_key, "")
                                    })
                                    added_count += 1
                    if added_count > 0:
                        st.success(f"✅ Added {added_count} personnel to manifest!")
                    else:
                        st.info("No new personnel to add.")
            
            # Display the manual input fields
            for i in range(1, 41):
                if st.session_state[f"p_show_{i}"]:
                    p_col1, p_col2, p_col3 = st.columns([2, 2, 2])
                    with p_col1:
                        p_name = st.text_input(f"Personnel Name {i}", key=f"name_{i}")
                    with p_col2:
                        p_comp = st.text_input(f"Company/Vendor {i}", key=f"comp_{i}")
                    with p_col3:
                        p_id = st.text_input(f"Security ID No. {i}", key=f"id_{i}")
                            
                    # Auto-add if name is entered and not already in list
                    if p_name and not any(p['name'] == p_name for p in st.session_state.personnel_list):
                        st.session_state.personnel_list.append({"name": p_name, "company": p_comp, "id_no": p_id})

            if visible_count < 40:
                if st.button("➕ Add More Personnel Fields", key="add_more_fields"):
                    for i in range(1, 41):
                        if not st.session_state[f"p_show_{i}"]:
                            st.session_state[f"p_show_{i}"] = True
                            st.rerun()
        
        if selection_method == "Mixed (Database + Manual)":
            st.subheader("Database Selection")
            if df_engr_tech_db is not None and not df_engr_tech_db.empty:
                with st.expander("📁 Select from Database", expanded=True):
                    # Get unique companies and regions
                    company_list = sorted(df_engr_tech_db['Company'].unique())
                    company_list = [c for c in company_list if str(c).strip() != '']
                    
                    region_list = []
                    if 'Region' in df_engr_tech_db.columns:
                        region_list = sorted(df_engr_tech_db['Region'].unique())
                        region_list = [r for r in region_list if str(r).strip() != '']
                        if 'N/A' in region_list:
                            region_list.remove('N/A')
                            region_list.append('N/A')
                        region_list = ['All Regions'] + region_list
                    else:
                        region_list = ['All Regions']
                    
                    # Filters
                    col_filters = st.columns(3)
                    
                    with col_filters[0]:
                        if company_list:
                            selected_companies = st.multiselect(
                                "Filter by Company:",
                                options=['All Companies'] + company_list,
                                default=['All Companies'],
                                key="company_filter_mixed"
                            )
                        else:
                            selected_companies = ['All Companies']
                    
                    with col_filters[1]:
                        if region_list:
                            selected_regions = st.multiselect(
                                "Filter by Region:",
                                options=region_list,
                                default=['All Regions'],
                                key="region_filter_mixed"
                            )
                        else:
                            selected_regions = ['All Regions']
                    
                    with col_filters[2]:
                        search_term = st.text_input("🔍 Search:", key="mixed_search")
                    
                    # Apply filters - FIXED LOGIC
                    filtered_engr = df_engr_tech_db.copy()
                    
                    # Company filter - FIXED: Only filter if not "All Companies"
                    if selected_companies and 'All Companies' not in selected_companies:
                        filtered_engr = filtered_engr[filtered_engr['Company'].isin(selected_companies)]
                    
                    # Region filter - FIXED: Only filter if not "All Regions"
                    if selected_regions and 'All Regions' not in selected_regions:
                        if 'Region' in filtered_engr.columns:
                            filtered_engr = filtered_engr[filtered_engr['Region'].isin(selected_regions)]
                    
                    if search_term:
                        filtered_engr = filtered_engr[
                            filtered_engr['Name'].astype(str).str.contains(search_term, case=False, na=False) |
                            filtered_engr['Company'].astype(str).str.contains(search_term, case=False, na=False) |
                            filtered_engr['ID No'].astype(str).str.contains(search_term, case=False, na=False)
                        ]
                    
                    if not filtered_engr.empty:
                        # Safe column selection
                        display_cols = []
                        if 'Name' in filtered_engr.columns:
                            display_cols.append('Name')
                        if 'Company' in filtered_engr.columns:
                            display_cols.append('Company')
                        if 'ID No' in filtered_engr.columns:
                            display_cols.append('ID No')
                        if 'Region' in filtered_engr.columns:
                            display_cols.append('Region')
                        
                        if display_cols:
                            st.dataframe(
                                filtered_engr[display_cols],
                                hide_index=True,
                                use_container_width=True
                            )
                        
                        options_list = filtered_engr.index.tolist()
                        
                        if options_list:
                            def format_person_mixed(row_idx):
                                try:
                                    name = filtered_engr.loc[row_idx, 'Name']
                                    company = filtered_engr.loc[row_idx, 'Company']
                                    return f"{name} - {company}"
                                except:
                                    return f"Person {row_idx}"
                            
                            selected_indices = st.multiselect(
                                "Select personnel:",
                                options=options_list,
                                format_func=format_person_mixed,
                                key="mixed_selection"
                            )
                            
                            col_add, col_add_all, col_spacer = st.columns([1, 1, 2])
                            
                            with col_add:
                                if st.button("➕ Add Selected", key="add_mixed"):
                                    if selected_indices:
                                        added_count = 0
                                        for idx in selected_indices:
                                            try:
                                                person = filtered_engr.loc[idx]
                                                if not any(p['name'] == person['Name'] for p in st.session_state.personnel_list):
                                                    st.session_state.personnel_list.append({
                                                        "name": person['Name'],
                                                        "company": person['Company'],
                                                        "id_no": person['ID No']
                                                    })
                                                    added_count += 1
                                            except:
                                                pass
                                        if added_count > 0:
                                            st.success(f"✅ Added {added_count} personnel!")
                                        else:
                                            st.info("All selected personnel are already in the manifest.")
                                    else:
                                        st.warning("Please select personnel first")
                            
                            with col_add_all:
                                if st.button("➕ Add All Visible", key="add_all_mixed"):
                                    all_indices = filtered_engr.index.tolist()
                                    if all_indices:
                                        added_count = 0
                                        for idx in all_indices:
                                            try:
                                                person = filtered_engr.loc[idx]
                                                if not any(p['name'] == person['Name'] for p in st.session_state.personnel_list):
                                                    st.session_state.personnel_list.append({
                                                        "name": person['Name'],
                                                        "company": person['Company'],
                                                        "id_no": person['ID No']
                                                    })
                                                    added_count += 1
                                            except:
                                                pass
                                        if added_count > 0:
                                            st.success(f"✅ Added all {added_count} visible personnel!")
                                        else:
                                            st.info("All visible personnel are already in the manifest.")
                                    else:
                                        st.warning("No personnel to add.")
            
            st.markdown("---")
            st.subheader("Manual Addition")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                manual_name = st.text_input("Name:", key="manual_name_mixed")
            with col2:
                manual_company = st.text_input("Company:", key="manual_company_mixed")
            with col3:
                manual_id = st.text_input("ID No.:", key="manual_id_mixed")
            
            if st.button("➕ Add Manual Entry", key="add_manual_mixed"):
                if manual_name:
                    if not any(p['name'] == manual_name for p in st.session_state.personnel_list):
                        st.session_state.personnel_list.append({
                            "name": manual_name,
                            "company": manual_company,
                            "id_no": manual_id
                        })
                        st.success(f"✅ Added {manual_name} to manifest!")
                    else:
                        st.warning(f"{manual_name} is already in the manifest.")
        
        # Display current personnel manifest
        if st.session_state.personnel_list:
            st.markdown("---")
            st.markdown(f"### 📋 Current Team Manifest ({len(st.session_state.personnel_list)} personnel)")
            
            manifest_df = pd.DataFrame(st.session_state.personnel_list)
            manifest_df['id_no'] = manifest_df['id_no'].apply(format_id_number)
            st.dataframe(manifest_df, hide_index=True, use_container_width=True)
            
            col_clear, col_spacer = st.columns([1, 3])
            with col_clear:
                if st.button("🗑️ Clear All Personnel", use_container_width=True):
                    st.session_state.personnel_list = []
                    st.rerun()
        else:
            st.info("No personnel added yet. Use the methods above to add team members.")
        
        st.markdown("---")

        # --- STEP 5: LOGIC EXECUTION & RENDER ---
        if st.button("🚀 Build Multi-Site RAAWA Spreadsheet", type="primary"):
            if not st.session_state.personnel_list:
                st.warning("Please add at least one personnel entry to populate the work manifest.")
            else:
                try:
                    # Clear previous files before generating new ones
                    st.session_state['generated_files'] = []
                    st.session_state['files_generated'] = False
                    
                    site_groups = split_sites_by_territory_and_fm(matching_sites)
                    
                    total_groups = len(site_groups)
                    
                    # Create a progress bar
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    file_details = []
                    total_files = total_groups
                    processed = 0
                    
                    for group_info in site_groups:
                        processed += 1
                        status_text.text(f"Processing file {processed} of {total_files}...")
                        progress_bar.progress(processed / total_files)
                        
                        group_sites = group_info['dataframe']
                        territory = group_info['territory']
                        facility_manager = group_info['facility_manager']
                        batch_num = group_info['batch_num']
                        total_batches = group_info['total_batches']
                        
                        group_region = selected_region
                        if selected_region == 'All Regions':
                            group_region = group_sites.iloc[0].get('REGION', 'MIN')
                            group_region = group_region if group_region in ['LUZ', 'VIS', 'MIN'] else 'MIN'
                        
                        # Determine which requisitioner to use
                        # If manual selection is set, use it; otherwise auto-determine
                        if selected_req_profile["name"] != "N/A" and st.session_state.req_selection_method != "Auto (Territory-based)":
                            # Use manually selected requisitioner
                            req_profile = selected_req_profile
                        else:
                            # Auto-determine based on territory, project, and region
                            req_profile = get_requisitioner_for_territory_and_project(
                                territory, 
                                selected_project, 
                                group_region
                            )
                        
                        buffer = create_raawa_file(
                            group_sites, 
                            st.session_state.personnel_list, 
                            scope_of_work, 
                            start_date, 
                            end_date, 
                            req_profile,
                            facility_manager,
                            batch_num,
                            total_batches
                        )
                        
                        if buffer is None:
                            st.error(f"Failed to create RAAWA file for Territory {territory}")
                            continue
                        
                        clean_fm = facility_manager.replace(" ", "_").replace("/", "_").replace(",", "")
                        if total_batches > 1:
                            filename = f"RAAWA_{selected_project}_{group_region}_Territory{territory}_{clean_fm}_Batch{batch_num}of{total_batches}_{group_sites.iloc[0]['PLAID']}.xlsx"
                            display_name = f"{selected_project} - {group_region} - Territory {territory} - {facility_manager[:25]} (Batch {batch_num}/{total_batches})"
                        else:
                            filename = f"RAAWA_{selected_project}_{group_region}_Territory{territory}_{clean_fm}_{group_sites.iloc[0]['PLAID']}.xlsx"
                            display_name = f"{selected_project} - {group_region} - Territory {territory} - {facility_manager[:30]}"
                        
                        file_details.append({
                            "name": filename,
                            "buffer": buffer,
                            "display_name": display_name,
                            "territory": territory,
                            "facility_manager": facility_manager,
                            "requisitioner": req_profile['name'],
                            "num_sites": len(group_sites),
                            "batch_info": f" (Batch {batch_num}/{total_batches})" if total_batches > 1 else ""
                        })
                        
                        st.success(f"✅ Created RAAWA #{len(file_details)}: {selected_project} - {group_region} - Territory {territory} - {facility_manager[:40]} ({len(group_sites)} sites){' - Batch ' + str(batch_num) + '/' + str(total_batches) if total_batches > 1 else ''} | Requisitioner: {req_profile['name']}")
                        
                        # Force garbage collection after each file
                        gc.collect()
                    
                    # Clear progress indicators
                    progress_bar.empty()
                    status_text.empty()
                    
                    # Store in session state
                    st.session_state['generated_files'] = file_details
                    st.session_state['files_generated'] = True
                    
                    # Show success message without rerun
                    if len(file_details) > 0:
                        st.success(f"✅ Successfully generated {len(file_details)} RAAWA files!")
                    else:
                        st.error("No files were generated. Please check the errors above.")
                        
                except Exception as ex:
                    st.error(f"Error: {ex}")
                    import traceback
                    st.error(traceback.format_exc())
        
        # --- DISPLAY DOWNLOAD BUTTONS FOR GENERATED FILES ---
        # This section displays without requiring a rerun
        if st.session_state.get('generated_files', []):
            st.markdown("---")
            st.markdown("## 📥 Download Generated RAAWA Files")
            
            file_details = st.session_state.get('generated_files', [])
            
            if len(file_details) == 1:
                st.success(f"🎉 Successfully generated 1 RAAWA file!")
                file_info = file_details[0]
                
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.download_button(
                        label=f"💾 Download {file_info['name']}",
                        data=file_info['buffer'],
                        file_name=file_info['name'],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="single_download",
                        use_container_width=True
                    )
                with col2:
                    st.caption(f"📋 {file_info['num_sites']} site(s){file_info['batch_info']}")
                
                st.markdown(f"""
                **Details:**
                - 📋 **Project:** {selected_project}
                - 🌍 **Region:** {selected_region if selected_region != 'All Regions' else file_info.get('region', 'MIN')}
                - 🏢 **Facility Manager:** {file_info['facility_manager']}
                - 👤 **Requisitioner:** {file_info['requisitioner']}
                - 📍 **Territory:** {file_info['territory']}
                """)
                
            else:
                st.success(f"🎉 Successfully generated {len(file_details)} RAAWA files!")
                st.markdown(f"**Project:** {selected_project} | **Region:** {selected_region if selected_region != 'All Regions' else 'Multiple'}")
                
                with st.expander("📁 View All Generated Files", expanded=True):
                    for idx, file_info in enumerate(file_details):
                        st.markdown(f"**File #{idx + 1}: {selected_project} - {selected_region if selected_region != 'All Regions' else 'Multiple'} - Territory {file_info['territory']}{file_info['batch_info']}**")
                        
                        col1, col2, col3 = st.columns([2, 1.5, 1.5])
                        
                        with col1:
                            st.download_button(
                                label=f"📄 {file_info['display_name'][:50]}",
                                data=file_info['buffer'],
                                file_name=file_info['name'],
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_{idx}_{file_info['territory']}_{file_info['facility_manager'][:20]}",
                                use_container_width=True
                            )
                        
                        with col2:
                            st.caption(f"🏢 FM: {file_info['facility_manager'][:35]}")
                            st.caption(f"📍 Territory: {file_info['territory']}")
                        
                        with col3:
                            st.caption(f"📋 Req: {file_info['requisitioner'][:30]}")
                            st.caption(f"🏗️ {file_info['num_sites']} site(s)")
                        
                        st.markdown("---")
            
            col_clear, col_spacer = st.columns([1, 3])
            with col_clear:
                if st.button("🔄 Clear Files & Start New RAAWA", use_container_width=True):
                    # Clear all session state related to files
                    st.session_state['generated_files'] = []
                    st.session_state['files_generated'] = False
                    gc.collect()
                    st.rerun()
    else:
        st.error("Failed to load databases. Please check that all required Excel files are present.")
